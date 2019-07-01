#!/usr/bin/env python3
"""
This module contains GUI elements and the main application for the aquami
package.
"""

import os
import sys

from sys import platform as sys_pf
if sys_pf == 'darwin':
    import matplotlib
    matplotlib.use("TkAgg")


import tkinter as tk
from tkinter import filedialog
from tkinter.scrolledtext import ScrolledText  
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.widgets import Slider, Button
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.backends.backend_pdf import PdfPages
from skimage.color import rgb2gray, gray2rgb
from skimage.restoration import denoise_bilateral
from scipy.misc import imread, imsave
import time
import pandas as pd

import copy

from aquami import tkSimpleDialog
from aquami import display
from aquami import measure
from aquami import segment
from aquami import inout
from aquami import autoSelect

__author__ = "Joshua Stuckner"
__version__ = '1.0'

class selectRect(object):
    """
    Draws a rectange on a figure and keeps track of the rectangle's size and
    location.  Used to select the image data and scale bar.

    Attributes
    ----------
    x0 : float64
        X coordinate (row) of start of rectangle.
    y0 : float 64
        Y coordinate (column) of start of rectangle.
    x1 : float64
        X coordinate (row) of end of rectangle.
    y1 : float 64
        Y coordinate (column) of end of rectangle.
    """
    
    def __init__(self):
        self.ax = plt.gca()
        self.rect = Rectangle((0,0), 1, 1,
                              facecolor='none',
                              edgecolor='#6CFF33',
                              linewidth=3)
        self.x0 = None
        self.y0 = None
        self.x1 = None
        self.y1 = None
        self.ax.add_patch(self.rect)
        self.ax.figure.canvas.mpl_connect('button_press_event', self.on_press)
        self.ax.figure.canvas.mpl_connect('button_release_event',
                                          self.on_release)

    def on_press(self, event):
        self.x0 = event.xdata
        self.y0 = event.ydata

    def on_release(self, event):
        self.x1 = event.xdata
        self.y1 = event.ydata
        if (self.x1 is not None and self.x0 is not None and
                            self.y1 is not None and self.y0 is not None):
            self.rect.set_width(self.x1 - self.x0)
            self.rect.set_height(self.y1 - self.y0)
            self.rect.set_xy((self.x0, self.y0))
            self.ax.figure.canvas.draw()

class inputScale(tkSimpleDialog.Dialog):        
    def body(self, master):

        self.varUnit = tk.StringVar()
        self.varUnit.set("nm")

        self.directions = tk.Label(master,
                text="Please manually input the scale bar number.")

        self.textInput = tk.Text(master, width=8, height=1)
        self.optUnit = tk.OptionMenu(master, self.varUnit, "nm", "μm", "mm")
        
        f = Figure()
        a = f.add_subplot(111)
        a.imshow(self.scale, cmap=plt.cm.gray, interpolation=None)
        canvas = FigureCanvasTkAgg(f, master=master)
        try:
            canvas.draw()
        except:
            canvas.show()

        
        canvas.get_tk_widget().grid(row=1, column=0, columnspan=2)
        canvas._tkcanvas.grid(row=1, column=0, columnspan=2)

        self.directions.grid(row=0, column=0, columnspan=2)

        self.textInput.grid(row=2, column=0)
        self.optUnit.grid(row=2, column=1)
 

        return self.textInput #initial focus

    def apply(self):
        mult = 1
        if self.varUnit.get() == "μm":
            mult = 1000
        elif self.varUnit.get() == "mm":
            mult = 1000000
        self.result = int(int(self.textInput.get("1.0",tk.END))*mult)


class manualThreshhold(object):
    def __init__(self, img, threshinit):
        self.img = img
        self.original = img.copy()
        self.img = denoise_bilateral(self.img,
                            sigma_color=0.05,  #increase this to blur more
                            sigma_spatial=1,
                            multichannel=False)
        self.fig, self.ax = plt.subplots()
        plt.subplots_adjust(left=0.01, bottom=0.20)
        self.manThresh = 0
        self.pltimage = self.ax.imshow(self.img, interpolation='nearest')
        self.ax_thresh = plt.axes([0.2, 0.1, 0.65, 0.03])
        self.s_thresh = Slider(self.ax_thresh, 'Threshold', 0, 255,
                               valinit=threshinit)
        self.s_thresh.on_changed(self.threshUpdate)
        self.threshUpdate(val=1)
        figManager = plt.get_current_fig_manager()
        try:
            figManager.window.showMaximized()
        except AttributeError: # TkAgg backend
            figManager.window.state('zoomed')
        plt.axis('off')       

    def threshUpdate(self, val):      
        val = self.s_thresh.val
        self.manThreshVal = self.s_thresh.val
        img = inout.uint8(self.img)
        self.mask = img >= int(val)
        overlay = display.overlayMask(self.original, self.mask, color='o',
                                      return_overlay=True)
        self.pltimage.set_data(overlay)        
        self.fig.canvas.draw_idle()

    def show(self):
        plt.ion()
        plt.show()
        while plt.get_fignums():
            try:
                plt.pause(0.1)
            except:
                pass
        plt.ioff()

    def getMask(self):
        return self.mask

        
def selectFile(title="Select image", initialdir=None, multiple=False):
    """
    Launches a dialog to select a file.

    Parameters
    ----------
    initialdir : str, optional
        The start path of the select a file dialog.

    multiple : bool, optional
        If true, allows the selecting of multiple files.
        
    Returns
    -------
    file : str
        The directory and name of the selected file.
    """
    file = filedialog.askopenfilename(
            initialdir=initialdir,
            multiple=multiple,
            title=title
            )
    return file


def selectFolder():
    """
    Launches a dialog to select a folder.

    Returns
    -------
    directory : str
        The directory of the selected folder.
    """   
    directory = filedialog.askdirectory(
            title='Select file'
            )
    return directory


def saveFile(ftype=None, title='Save file'): 
    file = filedialog.asksaveasfile(
            defaultextension=ftype,
            title=title
            )
    
    path = file.name
    file.close()
    return path


def manualSelectImage(img):
    '''
    Used to select the image data and the scale bar in a raw microscopy image
    that may contain other metadata besides the actual image.
    
    Parameters
    ----------
    img : ndarray
        Raw microscopy image.

    Returns
    -------
    im_data : ndarray
        Selected area containing just the image data

    scale : ndarray
        Selected area containing an image of the scale bar.
    '''

    # Get shape of image and convert to gray scale.
    try:
        rows, cols =  img.shape
    except ValueError: # Convert to gray scale
        img = rgb2gray(img)
        rows, cols =  img.shape
    
    # Select the image.
    plt.imshow(img, cmap=plt.get_cmap('gray'))
    plt.title('Select the image data and close this figure.')
    figManager = plt.get_current_fig_manager()
    try:
        figManager.window.showMaximized()
    except AttributeError: # TkAgg backend
        figManager.window.state('zoomed')
        
    a = selectRect()
    plt.axis('off')
    plt.ion()
    plt.show()
    while plt.get_fignums():
        try:
            plt.pause(0.1)
        except:
            pass
    plt.ioff() 
    # Make sure the selection isn't out of bounds or reversed.
    y0 = min(a.y0, a.y1)
    y1 = max(a.y0, a.y1)
    x0 = min(a.x0, a.x1)
    x1 = max(a.x0, a.x1)
    y0 = 0 if y0 < 0 else y0
    y1 = rows if y1 > rows else y1
    x0 = 0 if x0 < 0 else x0
    x1 = cols if x1 > cols else x1
    # Crop the raw image to the image data.
    im_data = img[int(y0):int(y1), int(x0):int(x1)]
    
    # Select the scale bar.
    plt.imshow(img, cmap=plt.get_cmap('gray'))
    plt.title('Draw a box around the scale bar and close this figure.' \
              ' No need to be precise as long as the whole bar is included.')
    figManager = plt.get_current_fig_manager()
    try:
        figManager.window.showMaximized()
    except AttributeError: # TkAgg backend
        figManager.window.state('zoomed')
        
    a = selectRect()
    plt.axis('off')
    plt.ion()
    plt.show()
    while plt.get_fignums():
        try:
            plt.pause(0.1)
        except:
            pass
    plt.ioff()
    # Make sure the selection isn't out of bounds or reversed.
    y0 = min(a.y0, a.y1)
    y1 = max(a.y0, a.y1)
    x0 = min(a.x0, a.x1)
    x1 = max(a.x0, a.x1)
    y0 = 0 if y0 < 0 else y0
    y1 = cols if y1 > cols else y1
    x0 = 0 if x0 < 0 else x0
    x1 = rows if x1 > rows else x1
    # Crop the raw image to the scale bar.   
    scale = img[int(a.y0):int(a.y1), int(a.x0):int(a.x1)]

    return im_data, scale


class CreateToolTip(object):
    '''
    create a tooltip for a given widget
    '''
    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.close)
        self.timer = []
        self.tw = None
        
    def enter(self, event=None):
        self.timer = self.widget.after(700, self.display)
        
    def display(self):
        x = y = 0
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        # creates a toplevel window
        self.tw = tk.Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tw, text=self.text, justify='left',
                       background='#ffffe6', relief='solid', borderwidth=1,
                       font=("times", "10", "normal"))
        label.pack(ipadx=5, ipady=3)
        
    def close(self, event=None):
        if self.tw is not None:
            self.tw.destroy()
        self.widget.after_cancel(self.timer)

class Aquami_Gui(tk.Frame):
    """
    This class in a GUI for BNMImage.
    """
    
    def __init__(self, master=None):
        """
        Called when class is created.
        """
        
        tk.Frame.__init__(self, master)
        self.master.title('AQUAMI')
        module_path, this_filename = os.path.split(__file__)
        try:
            self.master.iconbitmap(''.join((module_path, '/icon.ico')))
        except:
            pass

        self.initGUI()

    def initGUI(self):
        """
        Create and layout all the widgets.
        """

        self.pack(fill=tk.BOTH, expand=True)

        # Figure out sizing.
        width = 200
        height = 200
        pad = 5
        fontWidth = 8
        bigWidth = int((width*3 + pad*6) / fontWidth)
        
        # Create option frames.
        self.frameOptions = tk.LabelFrame(self, text="Options:",
                                          width=width, height=height)
        self.frameSegment = tk.LabelFrame(self, text="Segmentation Method:",
                                          width=width, height=height)
        self.frameMeasure = tk.LabelFrame(self, text="Measurements:",
                                          width=width, height=height)

        # Create text boxes and labels.
        self.labelStatus = tk.LabelFrame(self, text="Status:", bd=0)
        self.labelResults = tk.LabelFrame(self, text="Results:", bd=0)
        self.textStatus = ScrolledText(self.labelStatus, height=5,
                                       width=bigWidth)
        self.textResults = ScrolledText(self.labelResults, height=10,
                                        width=bigWidth)

        # Create buttons.
        self.buttonCalculate = tk.Button(self, text='Calculate',
                                         width=20, height=1, font=12, bd=3,
                                         command=lambda:self.prepare())
        self.buttonSaveAll = tk.Button(self, text='Save Session Summary',
                                       command=self.saveAll)
        self.buttonSelectOutFold = tk.Button(self, text='Set Output Folder',
                                             command=self.setOutputFolder)
        self.buttonAbout = tk.Button(self, text='About', command=self.about)

        # Arrange toplevel widgets.
        self.frameOptions.grid(row=0, column=2, padx=pad, pady=pad,
                               sticky='NESW')
        self.frameSegment.grid(row=0, column=1, padx=pad, pady=pad,
                               sticky='NESW')
        self.frameMeasure.grid(row=0, column=0, padx=pad, pady=pad,
                               sticky='NESW')

        self.buttonCalculate.grid(row=1, column=1, 
                                  padx=pad, pady=pad*3)
        self.buttonSelectOutFold.grid(row=1, column=0, 
                                  padx=pad, pady=pad*3)
        self.buttonAbout.grid(row=6, column=2, sticky='e', padx=20, pady=10)

        self.labelStatus.grid(row=2, column=0, columnspan=3, sticky='w',
                              padx=pad, pady=pad)
        self.textStatus.grid(row=3, column=0, columnspan=3)
        self.labelResults.grid(row=4, column=0, columnspan=3, sticky='w',
                               padx=pad, pady=pad)
        self.textResults.grid(row=5, column=0, columnspan=3)
        self.buttonSaveAll.grid(row=6, column=1, padx=pad, pady=pad)

        # Variables
        self.outFold = None
        columns = [["","","","",
                    "Bright phase diameter",
                    "","",
                    "Dark phase diameter",
                    "","",
                    "Bright length",
                    "","",
                    "Dark length",
                    "","",
                    "Bright area",
                    "","",
                    "Dark area",
                    "","",
                    "Bright connected length",
                    "","",
                    "Dark connected length",
                    "",""],                    
                   ["image",
                    "pixel size",
                    "area frac",
                    "est diam",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured",
                    "Average",
                    "SD",
                    "Measured"]]
                   
        self.saveAll = np.array(columns)

        # Measurement options.
        # Variables.
        self.varDiameter = tk.BooleanVar()
        self.varLength = tk.BooleanVar()
        self.varArea = tk.BooleanVar()
        self.varSumConnectedLength = tk.BooleanVar()
        self.varAreaFraction = tk.BooleanVar()
        # Create widgets.
        self.checkDiameter = tk.Checkbutton(self.frameMeasure,
                text="Diameter", variable=self.varDiameter)
        self.checkLength = tk.Checkbutton(self.frameMeasure,
                text="Length", variable=self.varLength)
        self.checkArea = tk.Checkbutton(self.frameMeasure,
                text="Area", variable=self.varArea)
        self.checkSumConnectedLength = tk.Checkbutton(self.frameMeasure,
                text="Connected length", variable=self.varSumConnectedLength)
        self.checkAreaFraction = tk.Checkbutton(self.frameMeasure,
                text="Area fraction", variable=self.varAreaFraction)
        # Pack widgets.
        self.checkDiameter.grid(row=0, column=0, sticky='w')
        self.checkLength.grid(row=1, column=0, sticky='w')
        self.checkArea.grid(row=2, column=0, sticky='w')
        self.checkSumConnectedLength.grid(row=3, column=0, sticky='w')
        self.checkAreaFraction.grid(row=4, column=0, sticky='w')
        # Check appropriate boxes.
        self.checkDiameter.select()
        self.checkLength.select()
        self.checkArea.select()
        self.checkSumConnectedLength.select()
        self.checkAreaFraction.select()
                
        # Segment options.
        # Variables.
        self.varSegment = tk.StringVar()
        # Create widgets.
        self.radAccurate = tk.Radiobutton(self.frameSegment,
                text="Accurate", variable=self.varSegment, value="accurate",
                command=self.updateOptions)
        self.radFast = tk.Radiobutton(self.frameSegment,
                text="Fast", variable=self.varSegment, value="fast",
                command=self.updateOptions)
        self.radManual= tk.Radiobutton(self.frameSegment,
                text="Manual", variable=self.varSegment, value="manual",
                command=self.updateOptions)
        self.radFromBinary = tk.Radiobutton(self.frameSegment,
                text="From binary", variable=self.varSegment, value="binary",
                command=self.updateOptions)
        # Pack widgets.
        self.radAccurate.grid(row=0, column=0, sticky='w')
        self.radFast.grid(row=1, column=0, sticky='w')
        self.radManual.grid(row=2, column=0, sticky='w')
        self.radFromBinary.grid(row=3, column=0, sticky='w')
        # Check appropriate boxes.
        self.radAccurate.select()

        # Option options.
        # Profiles
        profiles = autoSelect.profiles()
        # Variables.
        self.varShowSteps = tk.BooleanVar()
        self.varOutputExcel = tk.BooleanVar()
        self.varSavePDF = tk.BooleanVar()
        self.varSaveMovie = tk.BooleanVar()
        self.varSaveBinary = tk.BooleanVar()
        self.varAutoParse = tk.BooleanVar()
        self.varProfile = tk.StringVar()
        self.varProfile.set(profiles[0])
        # Create widgets.
        self.checkShowSteps = tk.Checkbutton(self.frameOptions,
                text="Show steps", variable=self.varShowSteps)
        self.checkOutputExcel = tk.Checkbutton(self.frameOptions,
                text="Output to Excel", variable=self.varOutputExcel)
        self.checkSavePDF = tk.Checkbutton(self.frameOptions,
                text="Save PDF", variable=self.varSavePDF)
        self.checkSaveMovie = tk.Checkbutton(self.frameOptions,
                text="Save movie", variable=self.varSaveMovie)
        self.checkSaveBinary = tk.Checkbutton(self.frameOptions,
                text="Save binary", variable=self.varSaveBinary)
        self.checkAutoParse = tk.Checkbutton(self.frameOptions,
                text="Auto parse raw image", variable=self.varAutoParse,
                command=self.updateAuto)
        self.optionProfile = tk.OptionMenu(self.frameOptions, self.varProfile,
                *profiles)
        self.optionProfile.config(state=tk.DISABLED)

        # Pack widgets.
        self.checkShowSteps.grid(row=0, column=0, sticky='w')
        self.checkOutputExcel.grid(row=1, column=0, sticky='w')
        self.checkSavePDF.grid(row=2, column=0, sticky='w')
        #self.checkSaveMovie.grid(row=3, column=0, sticky='w')
        self.checkSaveBinary.grid(row=4, column=0, sticky='w')
        self.checkAutoParse.grid(row=5, column=0, sticky='w')
        self.optionProfile.grid(row=6, column=0, sticky='w', padx=15)
        
        # Check appropriate boxes.
        self.checkOutputExcel.select()

        self.createToolTips()

    def createToolTips(self):
        self.ttps = []
        this_dir, this_filename = os.path.split(__file__)
        widgets = [[self.radAccurate, 'seg_accurate.txt'],
                   [self.radFast, 'seg_fast.txt'],
                   [self.radManual, 'seg_manual.txt'],
                   [self.radFromBinary, 'seg_binary.txt'],
                   [self.checkShowSteps, 'opt_show_steps.txt'],
                   [self.checkOutputExcel, 'opt_output_excel.txt'],
                   [self.checkSavePDF, 'opt_save_pdf.txt'],
                   [self.checkSaveMovie, 'opt_save_movie.txt'],
                   [self.checkSaveBinary, 'opt_save_binary.txt'],
                   [self.checkDiameter, 'mes_diameter.txt'],
                   [self.checkLength, 'mes_length.txt'],
                   [self.checkArea, 'mes_area.txt'],
                   [self.checkSumConnectedLength, 'mes_connected_length.txt'],
                   [self.checkAreaFraction, 'mes_area_fraction.txt'],
                   ]

        for widget, txt in widgets:
            text_path = os.path.join(this_dir, 'ttps', txt)
            f = open(text_path)
            data = f.read()
            f.close()
            self.ttps.append(CreateToolTip(widget, data))

    def setOutputFolder(self):
        self.outFold = selectFolder()
        
    def saveAll(self):
        """
        Saves the results of each image analyzed this session to an excel file.
        """
        path = saveFile(ftype='xlsx')
        writer = pd.ExcelWriter(path)
        df = pd.DataFrame(self.saveAll)
        df.to_excel(writer, header=False, index=False)
        writer.save()
        
        #Format the excel file
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side
            #Load the workbook and worksheet
            wb = openpyxl.load_workbook(filename=path)
            ws = wb.get_sheet_by_name("Sheet1")
            cells = ['E1','H1','K1','N1','Q1','T1','W1','Z1']
            ws.merge_cells('E1:G1')
            ws.merge_cells('H1:J1')
            ws.merge_cells('K1:M1')
            ws.merge_cells('N1:P1')
            ws.merge_cells('Q1:S1')
            ws.merge_cells('T1:V1')
            ws.merge_cells('W1:Y1')
            ws.merge_cells('Z1:AB1')
            #Bold and center the headers
            ft = Font(bold=True)
            for cell in cells:
                ws[cell].alignment = Alignment(horizontal="center")
                ws[cell].font = ft
            #Add borders
            rows,_ = self.saveAll.shape
            for i in range(rows):
                for cell in cells:
                    c = cell[0]+str(i+1)
                    ws[c].border = Border(left=Side(style='thin'))

                
            
            wb.save(path)
            
        except ImportError:
            pass

    def updateOptions(self):
        """
        Updates the GUI options.  Deactivates some options in response to the
        selection of other options.
        """
        if self.varSegment.get() == "binary":
            self.checkSaveBinary.config(state=tk.DISABLED)
        else:
            self.checkSaveBinary.config(state=tk.NORMAL)

    def updateAuto(self):
        """
        Makes the profile widget visible.
        """
        if self.varAutoParse.get():
            self.optionProfile.config(state=tk.NORMAL)
        else:
            self.optionProfile.config(state=tk.DISABLED)

    def write(self, box, *strings, sep=' ', end='\n'):
        """
        Displays text in a text box with similar functionality as the print
        statement.

        Parameters
        ----------
        box : tk.Text or ScrolledText
            Text box that will be written to.

        *strings : str (any number of)
            Strings that will be written in the text box.
        sep : str, optional
            Text that will be inserted between *strings arguements.
        end : str, optional
            Text that will be inserted at the end.
        """

        # Prepare the text to write.
        output = ''
        for i in range(len(strings)-1):            
            output = output+str(strings[i])+sep
        output = output+str(strings[-1])
        output = output + end

        # Write the output to the appropraite text box.
        box.insert(tk.END, output)
        box.see(tk.END)

        self.update_idletasks()
        time.sleep(0.1)

    def prepare(self):
        """
        Allows the selection of multiple files if segment is not set to binary.
        Calls the calculate function for each selected file.
        """
        if self.varSegment.get() == "binary":
            self.calculate("")
        else:
            files = selectFile(multiple=True)
            for file in files:
                self.calculate(file)

    def about(self):
        """
        Displays the about the software window.
        """
        
        try:
            if self.tlAbout.winfo_exists():
                exists = True
            else:
                exists = False
        except AttributeError:
            exists = False

        if not exists:    
            self.tlAbout = tk.Toplevel(self)
            self.tlAbout.title('About AQUAMI')
            module_path, this_filename = os.path.split(__file__)
            try:
                self.tlAbout.iconbitmap(''.join((module_path, '/icon.ico')))
            except:
                pass
            self.buttonAboutClose = tk.Button(self.tlAbout, text='OK', width=8,
                                              command=self.aboutClose)
            self.frameAbout = tk.Frame(self.tlAbout)
            self.frameAbout.grid(row=0, column=0, padx=5, pady=5, ipadx=0, ipady=0)
            self.buttonAboutClose.grid(row=1, column=0, sticky='e', pady=5, padx=15)

            self.aText = tk.Text(self.frameAbout, background='SystemButtonFace',
                                 width=75, height=18, font="Times 11", bd=0,
                                 exportselection=False, cursor='arrow')
            self.aText.pack(anchor='w')

            self.aText.insert(tk.INSERT, "AQUAMI\n", )
            self.aText.tag_add('title', '1.0', '1.20')
            self.aText.tag_config('title', font='Times 16 bold')
            self.aText.insert(tk.END, ''.join((
                'Version: ', __version__, '\n', 
                'Author: Joshua Stuckner\n',
                'Contact: stuckner@vt.edu\n',
                'Liscence: MIT\n\n',
                'This software was developed by a graduate student at Virginia Tech and released in\n',
                'the hopes that it will be useful. This software may be freely modified and/or distributed. \n',
                'The author plans to support this software until at least Dec. 2018. Please contact the\n',
                'author to report any issues or bugs, or to request additional features.  Highly demanded\n',
                'features may be added to the software.\n\n',
                'If this software is useful to your research, please consider citing:\n',
                'J. A. Stuckner et al., "AQUAMI: An Open Source Python Package and GUI for the\n',
                'Automatic Quantitative Analysis of Morphologically Complex Multiphase Materials,"\n',
                'Computational Materials Science, vol. 139, pp. 320-329, Nov. 2017\n',
                'DOI: 10.1016/j.commatsci.2017.08.012'
                )))

            self.aText.tag_add('citing', '13.0', '13.100')
            self.aText.tag_config('citing', font='Times 11 bold')
            self.aText.tag_add('etal', '14.15', '14.21')
            self.aText.tag_config('etal', font='Times 11 italic')
            self.aText.tag_add('journal', '16.0', '16.31')
            self.aText.tag_config('journal', font='Times 11 italic')
            self.aText.tag_add('doi', '17.0', '17.111')
            self.aText.tag_config('doi', foreground='blue', underline=1)
            self.aText.tag_bind('doi', '<Button-1>', self.hyperlink_click)
            self.aText.tag_bind("doi", "<Enter>", self.hyper_enter)
            self.aText.tag_bind("doi", "<Leave>", self.hyper_leave)
            
            self.aText.config(state=tk.DISABLED)

    def hyperlink_click(self, event):
        url = r'https://doi.org/10.1016/j.commatsci.2017.08.012'
        if sys.platform=='win32':
            os.startfile(url)
        elif sys.platform=='darwin':
            subprocess.Popen(['open', url])
        else:
            try:
                subprocess.Popen(['xdg-open', url])
            except OSError:
                print(''.join(('Please open a browser on: ', url)))

    def hyper_enter(self, event):
            self.aText.config(cursor="hand2")

    def hyper_leave(self, event):
        self.aText.config(cursor="arrow")

    def aboutClose(self):
        """
        Closes the about the software window
        """
        self.tlAbout.destroy()
        

    def calculate(self, fullFile):
        """
        Performs measurements based on the options.

        Parameters
        ----------
        fullFile : str
            The path to the image to operate on.
        """

        # Initialize variables
        pixelSize = 0
        areaFracBright = 0
        estSize= 0
        diameter_average_bright = 0
        diameter_SD_bright = 0
        diameter_num_bright = 0
        diameter_average_dark = 0
        diameter_SD_dark = 0
        diameter_num_dark = 0
        length_average_bright = 0
        length_SD_bright = 0
        length_num_bright = 0
        length_average_dark = 0
        length_SD_dark = 0
        length_num_dark = 0
        area_average_bright = 0
        area_SD_bright = 0
        area_num_bright = 0
        area_average_dark = 0
        area_SD_dark = 0
        area_num_dark = 0
        sumLength_average_bright = 0
        sumLength_SD_bright = 0
        sumLength_num_bright = 0
        sumLength_average_dark = 0
        sumLength_SD_dark = 0
        sumLength_num_dark = 0
        
        # Rename text boxes and tkinter variables for easier reference
        results = self.textResults
        status = self.textStatus
        showSteps = self.varShowSteps.get()
        outputExcel = self.varOutputExcel.get()
        savePDF = self.varSavePDF.get()

        # Handle strange GUI input.
        if self.varSegment.get() == "binary":
            self.varSaveBinary.set(False)

        # Select the image to operate on.
        if self.varSegment.get() == "binary":
            brightFile = selectFile(title="Select the bright phase mask")
            darkFile = selectFile(title="Select the bright phase mask")
            folder = '/'.join(brightFile.split('/')[:-1])
            fname = brightFile.split('/')[-1].split(' -')[:-1][0]
            ftype = brightFile.split('.')[-1]
            fnametype = fname + '.' + ftype
            pixelString = brightFile.split('--')[1]            
            pixelSize = float(pixelString)
            fullFile = brightFile

        else:
            #fullFile = selectFile()
            # Figure out file name and type.
            folder = '/'.join(fullFile.split('/')[:-1])
            fname = fullFile.split('/')[-1].split('.')[:-1][0]
            ftype = fullFile.split('.')[-1]
            fnametype = fname + '.' + ftype

        # Status update.
        if self.varSegment.get() == "binary":
            self.write(status, "Operating on", fnametype, "(binary)")
            self.write(results, "Operating on", fnametype, "(binary)")
        else:
            self.write(status, "Operating on", fnametype)
            self.write(results, "Operating on", fnametype)

        # Load image.
        img_raw = inout.load(fullFile)

        # Prepare for output of data to excel
        if outputExcel:
            if self.outFold == None:
                excelPath = folder+'/'+fname+' - All Measurements.xlsx'
                writer = pd.ExcelWriter(excelPath)
            else:
                excelPath = self.outFold+'/'+fname+' - All Measurements.xlsx'
                writer = pd.ExcelWriter(excelPath)

        # Prepare PDF file.
        pdf = None
        if savePDF:
            if self.outFold == None:
                pdf = PdfPages(folder+'/'+fname+' - Steps.pdf')
            else:
                pdf = PdfPages(self.outFold+'/'+fname+' - Steps.pdf')

        # Show steps.
        if showSteps:
            display.showFull(img_raw, title="Raw Image.")

        if savePDF:
            inout.pdfSaveImage(pdf, img_raw, title="Raw Image.")

        # Select the image data and scale bar.
        if self.varSegment.get() != "binary":

            if self.varAutoParse.get():
                profile = self.varProfile.get()
                img, pixelSize = autoSelect.autoDetect(img_raw, profile=profile)
            else:
                img, scale = manualSelectImage(img_raw)
                
                # Ensure proper dtype.
                img = inout.uint8(img) 
                scale = inout.uint8(scale)

                # Get the pixel size.
                pixelSize = measure.manualPixelSize(scale)
                self.write(results, "The pixel size is %.3f nm" %(pixelSize))
            
            # Get a rough estimated of the ligament diameter.
            bright, dark = segment.roughSegment(img)
            estSizeD = measure.estimateDiameter(dark)
            estSizeB = measure.estimateDiameter(bright)
            estSizeB_rough = copy.copy(estSizeB)
            estSizeD_rough = copy.copy(estSizeD)
            # Get the average estimated size with a weight towards the thinner
            estSize = (min(estSizeD, estSizeB) + estSizeD + estSizeB) / 3
            # Result update                
            self.write(results, "Weighted average estimated diameter: ",
                       round(estSize,3)," pixels (", round(estSize*pixelSize,3),
                       " nm).", sep='')
            if savePDF:
                inout.pdfSaveImage(pdf, img, title="Selected image data",
                                   cmap=plt.cm.gray)
        else:
            img = img_raw.copy()
            img = inout.uint8(img)

        # Segment the image.
        # Status update.
        if self.varSegment.get() == "accurate":
            self.write(status, "Segmenting image. This may take a while,",
                       "especially at higher resolutions.")
        else:
            self.write(status, "Segmenting image.")
        # Segment based on options.    
        if self.varSegment.get() == "fast":
            #already did this when estimating size
            pass
        elif self.varSegment.get() == "manual":
            bright, dark = segment.manualSegment(img)
        elif self.varSegment.get() == "accurate":
            bright, dark = segment.segment(img, estSize)
        elif self.varSegment.get() == "binary":
            bright = imread(brightFile)
            dark = imread(darkFile)
            estSizeD = measure.estimateDiameter(dark)
            estSizeB = measure.estimateDiameter(bright)
            estSizeB_rough = copy.copy(estSizeB)
            estSizeD_rough = copy.copy(estSizeD)
            # Get the average estimated size with a weight towards the thinner
            estSize = (min(estSizeD, estSizeB) + estSizeD + estSizeB) / 3
            # Result update                
            self.write(results, "Weighted average estimated diameter: ",
                       round(estSize,3)," pixels (", round(estSize*pixelSize,3),
                       " nm).", sep='')
            bright = inout.uint8(bright)
            dark = inout.uint8(dark)

        if showSteps and self.varSegment.get() != "binary":
            display.overlayMask(img, bright, title="Bright phase mask.",
                                animate=True)
            display.overlayMask(img, dark, title="Dark phase mask.",
                                animate=True)

        if savePDF and self.varSegment.get() != "binary":
            inout.pdfSaveOverlay(pdf, img, bright, title="Bright phase mask.")
            inout.pdfSaveOverlay(pdf, img, dark, title="Dark phase mask.")

        if self.varSaveMovie.get():
            stack = []
            for i in range(5):
                stack.append(gray2rgb(img))
                stack.append(display.overlayMask(img,bright,
                                                 return_overlay=True))
            inout.save_movie(stack, ''.join((folder, '//', fname,
                                             ' - segment movie.mp4')))

        if self.varSaveBinary.get():
            inout.saveBinary(bright,
                folder+'/'+fname+' - Bright Segment--'+str(pixelSize)+'--.tif')
            inout.saveBinary(dark,
                folder+'/'+fname+' - Dark Segment--'+str(pixelSize)+'--.tif')

        
              
        # Calculate area fraction.
        if self.varAreaFraction.get():
            # Status update.
            self.write(status, "Calculating area fraction.")
            areaFracBright = np.count_nonzero(bright) /        \
                    (np.count_nonzero(bright) + np.count_nonzero(dark))
            # Results update.
            self.write(results, "Bright phase area fraction: %.3f" \
                       %(areaFracBright))

        # Refine the estimated diameters.
        estSizeB, _ = measure.diameter(bright, estSize)
        estSizeD, _ = measure.diameter(dark, estSize)

                           
        # Calculate the diameter.
        if self.varDiameter.get():
            # Status update.
            self.write(status, "Calculating ligament diameter.")
            # Calculate.
            diameter_average_bright, diameter_SD_bright, diameter_all_bright =\
                            measure.diameter(bright, estSizeB,
                                             showSteps=showSteps,
                                             returnAll=True,
                                             pdf=pdf)
            diameter_num_bright = len(diameter_all_bright)
            diameter_average_dark, diameter_SD_dark, diameter_all_dark =\
                            measure.diameter(dark, estSizeD,
                                             showSteps=showSteps,
                                             returnAll=True,
                                             pdf=pdf)
            diameter_num_dark = len(diameter_all_dark)
            # Update estimated ligament sizes.
            estSizeB = diameter_average_bright
            estSizeD = diameter_average_dark
            # Results update.
            self.write(results, "Bright phase diameter:",
                       round(diameter_average_bright,2), "±",
                       round(diameter_SD_bright), "pixels.")
            self.write(results, "Dark phase diameter:",
                       round(diameter_average_dark,2), "±",
                       round(diameter_SD_dark), "pixels.")

            if outputExcel:
                df = pd.DataFrame(diameter_all_bright)
                df.to_excel(writer, sheet_name="Bright phase diameter data")
                df = pd.DataFrame(diameter_all_dark)
                df.to_excel(writer, sheet_name="Dark phase diameter data") 
                
        # Fix estSizeB and estSizeD if there is a problem
        estSizeB = estSizeB_rough if np.isnan(estSizeB) else estSizeB      
        estSizeD = estSizeD_rough if np.isnan(estSizeD) else estSizeD         
            
        # Calculate the ligament length.
        if self.varLength.get():
            # Status update.
            self.write(status, "Calculating ligament length.")
            length_average_bright, length_SD_bright, length_all_bright = \
                            measure.length(bright, estSizeB,
                                           showSteps=showSteps,
                                           returnAll=True,
                                           pdf=pdf)
            try:
                length_num_bright = len(length_all_bright)
            except TypeError:
                length_num_bright = 0
            length_average_dark, length_SD_dark, length_all_dark = \
                            measure.length(dark, estSizeD,
                                           showSteps=showSteps,
                                           returnAll=True,
                                           pdf=pdf)
            try:
                length_num_dark = len(length_all_dark)
            except TypeError:
                length_num_dark = 0
            # Results update.
            if length_average_bright != 0:
                self.write(results, "Bright phase ligament length:",
                           round(length_average_bright,2), "±",
                           round(length_SD_bright), "pixels.")
            if length_average_dark != 0:
                self.write(results, "Dark phase ligament length:",
                           round(length_average_dark,2), "±",
                           round(length_SD_dark), "pixels.")

            if outputExcel:
                if length_average_bright != 0:
                    df = pd.DataFrame(length_all_bright)
                    df.to_excel(writer, sheet_name="Bright phase length data")
                if length_average_dark != 0:
                    df = pd.DataFrame(length_all_dark)
                    df.to_excel(writer, sheet_name="Dark phase length data") 

        # Calculate average object area.
        if self.varArea.get():
            # Status update.
            self.write(status, "Calculating average object area.")
            area_average_bright, area_SD_bright, area_all_bright = \
                            measure.area(bright, estSizeB,
                                           showSteps=showSteps,
                                           returnAll=True)
            try:
                area_num_bright = len(area_all_bright)
            except TypeError:
                area_num_bright = 0
            area_average_dark, area_SD_dark, area_all_dark = \
                            measure.area(dark, estSizeD,
                                           showSteps=showSteps,
                                           returnAll=True)
            try:
                area_num_dark = len(area_all_dark)
            except TypeError:
                area_num_dark = 0
            # Results update.
            if area_average_bright != 0:
                self.write(results, "Bright phase average object area:",
                           round(area_average_bright,2), "±",
                           round(area_SD_bright), "pixels^2.")
            if area_average_dark != 0:
                self.write(results, "Dark phase average object area:",
                           round(area_average_dark,2), "±",
                           round(area_SD_dark), "pixels^2.")

            if outputExcel:
                if area_average_bright != 0:
                    df = pd.DataFrame(area_all_bright)
                    df.to_excel(writer, sheet_name="Bright phase area data")
                if area_average_dark != 0:
                    df = pd.DataFrame(area_all_dark)
                    df.to_excel(writer, sheet_name="Dark phase area data") 

        # Calculate the sum connected length of each object.
        if self.varSumConnectedLength.get():
            # Status update.
            self.write(status, "Calculating sum connected ligament length.")
            sumLength_average_bright, sumLength_SD_bright, \
                                      sumLength_all_bright = \
                            measure.connectedLength(bright, estSizeB,
                                        showSteps=showSteps,
                                        returnAll=True,
                                        pdf=pdf)
            try:
                sumLength_num_bright = len(sumLength_all_bright)
            except TypeError:
                sumLength_num_bright = 0
            sumLength_average_dark, sumLength_SD_dark, sumLength_all_dark = \
                            measure.connectedLength(dark, estSizeD,
                                        showSteps=showSteps,
                                        returnAll=True,
                                        pdf=pdf)
            try:
                sumLength_num_dark = len(sumLength_all_dark)
            except TypeError:
                sumLength_num_dark = 0
            # Results update.
            if sumLength_average_bright != 0:
                self.write(results, "Bright phase sum connected length:",
                           round(sumLength_average_bright,2), "±",
                           round(sumLength_SD_bright), "pixels.")
            if sumLength_average_dark != 0:
                self.write(results, "Dark phase sum connected length:",
                           round(sumLength_average_dark,2), "±",
                           round(sumLength_SD_dark), "pixels.")

            if outputExcel:
                if sumLength_average_bright != 0:
                    df = pd.DataFrame(sumLength_all_bright)
                    df.to_excel(writer,
                        sheet_name="Bright phase sum length data")
                if sumLength_average_dark != 0:
                    df = pd.DataFrame(sumLength_all_dark)
                    df.to_excel(writer,
                        sheet_name="Dark phase sum length data") 

        # Save excel file
        if outputExcel:
            writer.save()

        # Save the pdf.
        if savePDF:
            plt.title('Summary')
            plt.axis('off')
            fs = 10
            yloc = 0.95
            xloc = 0.01
            space = 0.05
            out = 'Image: ' + fnametype
            plt.text(xloc, yloc, out, fontsize=fs)
            yloc -= space
            out = 'Pixel size: ' + str(pixelSize)
            plt.text(xloc, yloc, out, fontsize=fs)
            yloc -= space
            out = 'Bright phase area fraction: ' + str(round(areaFracBright,3))
            plt.text(xloc, yloc, out, fontsize=fs)
            yloc -= space
            out = 'Rough diameter estimate: ' + str(round(estSize,3))+' pixels.'
            plt.text(xloc, yloc, out, fontsize=fs)
            yloc -= space*2

            out = 'Bright phase:'
            plt.text(xloc, yloc, out, fontsize=fs, fontweight='bold')
            yloc -= space
            if diameter_average_bright != 0:
                out = ('Ligament diameter: ' +
                       str(round(diameter_average_bright*pixelSize,3)) +
                       ' ± ' + str(round(diameter_SD_bright*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if length_average_bright != 0:
                out = ('Ligament length: ' +
                       str(round(length_average_bright*pixelSize,3)) +
                       ' ± ' + str(round(length_SD_bright*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if area_average_bright != 0:
                out = ('Object area: ' +
                       str(round(area_average_bright*pixelSize**2,3)) +
                       ' ± ' + str(round(area_SD_bright*pixelSize**2,3)) +
                       r' nm$^2$')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if sumLength_average_bright != 0:
                out = ('Sum connected length: ' +
                       str(round(sumLength_average_bright*pixelSize,3)) +
                       ' ± ' + str(round(sumLength_SD_bright*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            yloc -= space

            out = 'Dark phase:'
            plt.text(xloc, yloc, out, fontsize=fs, fontweight='bold')
            yloc -= space
            if diameter_average_dark != 0:
                out = ('Ligament diameter: ' +
                       str(round(diameter_average_dark*pixelSize,3)) +
                       ' ± ' + str(round(diameter_SD_dark*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if length_average_dark != 0:
                out = ('Ligament length: ' +
                       str(round(length_average_dark*pixelSize,3)) +
                       ' ± ' + str(round(length_SD_dark*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if area_average_dark != 0:
                out = ('Object area: ' +
                       str(round(area_average_dark*pixelSize**2,3)) +
                       ' ± ' + str(round(area_SD_dark*pixelSize**2,3)) +
                       r' nm$^2$')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            if sumLength_average_dark != 0:
                out = ('Sum connected length: ' +
                       str(round(sumLength_average_dark*pixelSize,3)) +
                       ' ± ' + str(round(sumLength_SD_dark*pixelSize,3)) +
                       ' nm')
                plt.text(xloc, yloc, out, fontsize=fs)
                yloc -= space
            

            pdf.savefig()
            plt.close()
            pdf.close()

        # Append data to the saveAll
        newdata = np.array(
                [[fnametype,
                  round(pixelSize,2),
                  round(areaFracBright,3),
                  round(estSize*pixelSize,3),
                  round(diameter_average_bright*pixelSize,3),
                  round(diameter_SD_bright*pixelSize,3),
                  diameter_num_bright,
                  round(diameter_average_dark*pixelSize,3),
                  round(diameter_SD_dark*pixelSize,3),
                  diameter_num_dark,
                  round(length_average_bright*pixelSize,3),
                  round(length_SD_bright*pixelSize,3),
                  length_num_bright,
                  round(length_average_dark*pixelSize,3),
                  round(length_SD_dark*pixelSize,3),
                  length_num_dark,
                  round(area_average_bright*pixelSize**2,3),
                  round(area_SD_bright*pixelSize**2,3),
                  area_num_bright,
                  round(area_average_dark*pixelSize**2,3),
                  round(area_SD_dark*pixelSize**2,3),
                  area_num_dark,
                  round(sumLength_average_bright*pixelSize,3),
                  round(sumLength_SD_bright*pixelSize,3),
                  sumLength_num_bright,
                  round(sumLength_average_dark*pixelSize,3),
                  round(sumLength_SD_dark*pixelSize,3),
                  sumLength_num_dark
                  ]])
        

        
        self.saveAll = np.append(self.saveAll, newdata, axis=0)

        # Add new line to status and results
        self.write(results, '')
        self.write(status, '')
        
if __name__ == "__main__":
    myapp = Aquami_Gui()
    myapp.mainloop()

def run():
    myapp = Aquami_Gui()
    myapp.mainloop()

